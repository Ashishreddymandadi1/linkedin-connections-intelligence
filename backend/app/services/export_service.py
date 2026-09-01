"""Export enriched profiles to an .xlsx workbook.

Four sheets: Profiles (one row per person), Experiences, Education, Skills.
Only enriched people (READY / PARTIAL / WAITING_FOR_FREE_LLM) are included.
"""
from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import repositories as repo
from app.constants import EnrichmentState
from app.models import Dataset

#: XML 1.0 forbids these control chars in a worksheet cell (openpyxl raises
#: IllegalCharacterError). Keeps normal unicode, newlines, tabs, bullets,
#: em-dashes, accents — strips only the truly illegal range (spec §32).
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def sanitize_excel_value(value):
    """Make any value safe to write into an openpyxl cell."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_RE.sub("", value)
    return value


def _append(ws, row: list) -> None:
    """Every row goes through here — sanitize each cell (spec §32: apply to
    EVERY value, not just experience.description)."""
    ws.append([sanitize_excel_value(v) for v in row])

_INCLUDED = {
    EnrichmentState.READY,
    EnrichmentState.PARTIAL,
    EnrichmentState.WAITING_FOR_FREE_LLM,
    EnrichmentState.LLM_COMPLETE,
    EnrichmentState.NORMALIZED,
}

_HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _sheet(wb: Workbook, title: str, headers: list[str]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    ws.freeze_panes = "A2"
    return ws


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(60, max(12, width + 2))


def _join(parts: list[str], sep: str = " | ") -> str:
    return sep.join(p for p in parts if p)


def build_workbook(db: Session, dataset: Dataset) -> bytes:
    people = [p for p in repo.list_people(db, dataset.id) if p.enrichment_state in _INCLUDED]

    wb = Workbook()
    wb.remove(wb.active)

    prof = _sheet(
        wb,
        "Profiles",
        [
            "Name", "First Name", "Last Name", "LinkedIn URL", "Public ID",
            "Headline", "Current Title", "Current Company", "Location",
            "City", "State", "Country", "Connections", "Followers",
            "Open To Work", "Hiring", "Premium", "Verified",
            "Data Confidence", "State",
            "Seniority", "Job Families", "Technical Domains", "Industries",
            "Years Experience", "Career Summary",
            "Explicit Skills", "Inferred Skills",
            "# Experiences", "# Education", "# Skills",
            "Experience Summary", "Education Summary",
            "CSV Company", "CSV Position", "Connected On", "Last Scraped",
        ],
    )
    exp_ws = _sheet(
        wb, "Experiences",
        ["Person", "LinkedIn URL", "Position", "Company", "Company URL",
         "Employment Type", "Location", "Start", "End", "Is Current", "Duration", "Description"],
    )
    edu_ws = _sheet(
        wb, "Education",
        ["Person", "LinkedIn URL", "School", "Degree", "Field of Study", "Start Year", "End Year"],
    )
    skill_ws = _sheet(
        wb, "Skills",
        ["Person", "LinkedIn URL", "Skill", "Source", "Inferred", "Confidence", "Evidence"],
    )
    cert_ws = _sheet(
        wb, "Certifications",
        ["Person", "LinkedIn URL", "Certification", "Issuer", "Issued", "URL"],
    )
    lang_ws = _sheet(wb, "Languages", ["Person", "LinkedIn URL", "Language", "Proficiency"])
    pub_ws = _sheet(
        wb, "Publications",
        ["Person", "LinkedIn URL", "Title", "Publisher", "Published", "URL"],
    )

    for p in people:
        exps = repo.get_experiences(db, p.id)
        edus = repo.get_education(db, p.id)
        skills = repo.get_skills(db, p.id)
        sem = repo.get_semantic(db, p.id)
        s = sem.data if sem and sem.data else {}

        exp_summary = _join(
            [
                f"{e.position or '?'} @ {e.company_name or '?'}"
                + (f" ({e.start_year}–{e.end_year or 'present'})" if e.start_year else "")
                for e in exps[:8]
            ]
        )
        edu_summary = _join(
            [
                _join([e.degree or "", e.field_of_study or "", e.school_name or ""], ", ")
                for e in edus[:5]
            ]
        )
        inferred = _join(
            [f"{i.get('skill')} ({i.get('confidence')})" for i in s.get("inferred_skills", []) if isinstance(i, dict)],
            ", ",
        )

        _append(prof, 
            [
                p.full_name, p.first_name, p.last_name, p.linkedin_url, p.public_identifier,
                p.headline, p.current_title, p.current_company, p.location_text,
                p.city, p.state, p.country, p.connections_count, p.followers_count,
                _yn(p.open_to_work), _yn(p.hiring), _yn(p.premium), _yn(p.verified),
                p.profile_completeness, p.enrichment_state,
                s.get("seniority_level"),
                ", ".join(s.get("job_families", [])),
                ", ".join(s.get("technical_domains", [])),
                ", ".join(s.get("industries", [])),
                s.get("years_of_experience"),
                s.get("career_summary"),
                ", ".join(s.get("explicit_skills", [])) or ", ".join(sk.skill_name for sk in skills if not sk.is_inferred),
                inferred,
                len(exps), len(edus), len(skills),
                exp_summary, edu_summary,
                p.csv_company, p.csv_position, p.connected_on,
                _dt(p.last_scraped_at),
            ]
        )

        for e in exps:
            _append(exp_ws, [
                p.full_name, p.linkedin_url, e.position, e.company_name, e.company_linkedin_url,
                e.employment_type, e.location,
                _period(e.start_month, e.start_year, e.start_text),
                "Present" if e.is_current else _period(e.end_month, e.end_year, e.end_text),
                _yn(e.is_current), e.duration_text, e.description,
            ])
        for e in edus:
            _append(edu_ws, [p.full_name, p.linkedin_url, e.school_name, e.degree, e.field_of_study, e.start_year, e.end_year])
        for sk in skills:
            _append(skill_ws, [
                p.full_name, p.linkedin_url, sk.skill_name, sk.source,
                _yn(sk.is_inferred), round(sk.confidence, 2), sk.evidence,
            ])
        for c in repo.get_certifications(db, p.id):
            _append(cert_ws, [p.full_name, p.linkedin_url, c.name, c.issuer, c.issued_at, c.url])
        for lang in repo.get_languages(db, p.id):
            _append(lang_ws, [p.full_name, p.linkedin_url, lang.name, lang.proficiency])
        for pub in repo.get_publications(db, p.id):
            _append(pub_ws, [p.full_name, p.linkedin_url, pub.title, pub.publisher, pub.published_at, pub.url])

    for ws in wb.worksheets:
        _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _yn(v) -> str:
    if v is None:
        return ""
    return "Yes" if v else "No"


def _dt(v: datetime | None) -> str:
    return v.strftime("%Y-%m-%d %H:%M") if v else ""


def _period(month: int | None, year: int | None, text: str | None) -> str:
    if text:
        return text
    if year and month:
        return f"{month:02d}/{year}"
    return str(year) if year else ""

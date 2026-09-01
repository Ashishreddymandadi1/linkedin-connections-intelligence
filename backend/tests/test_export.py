from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook

FIXTURE_CSV = Path(__file__).resolve().parents[1] / "fixtures" / "connections_sample.csv"


def test_export_xlsx_has_four_sheets_and_profile_rows(client):
    ds_id = client.post("/datasets", files={"file": ("c.csv", FIXTURE_CSV.read_bytes(), "text/csv")}).json()[
        "dataset"
    ]["dataset_id"]
    client.post(f"/datasets/{ds_id}/enrich")

    r = client.get(f"/datasets/{ds_id}/export")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in r.headers["content-disposition"]
    assert r.headers["content-disposition"].endswith('.xlsx"')

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == [
        "Profiles", "Experiences", "Education", "Skills",
        "Certifications", "Languages", "Publications",
    ]

    prof = wb["Profiles"]
    assert prof.max_row >= 6  # 5 fixtures + synthesized
    headers = [c.value for c in prof[1]]
    assert "LinkedIn URL" in headers and "Data Confidence" in headers

    names = {prof.cell(row=r_, column=1).value for r_ in range(2, prof.max_row + 1)}
    assert "Jane Smith" in names

    exp = wb["Experiences"]
    assert exp.max_row > 1
    assert any(
        exp.cell(row=r_, column=4).value == "Amazon" for r_ in range(2, exp.max_row + 1)
    )


def test_export_404_for_unknown_dataset(client):
    assert client.get("/datasets/nope/export").status_code == 404


def test_sanitize_strips_only_illegal_control_chars():
    from app.services.export_service import sanitize_excel_value

    dirty = "Led \x00team\x01 of \x0b6\x0c engineers\x1f — with • bullets, é accents,\nnewlines\tand tabs"
    clean = sanitize_excel_value(dirty)
    assert "\x00" not in clean and "\x01" not in clean and "\x0b" not in clean and "\x1f" not in clean
    # normal content is preserved
    assert "•" in clean and "é" in clean and "\n" in clean and "\t" in clean and "—" in clean
    assert "Led team of 6 engineers" in clean
    assert sanitize_excel_value(42) == 42 and sanitize_excel_value(None) is None


def test_export_survives_illegal_chars_in_profile_fields(client, db):
    """spec §34 TEST 10 — control chars in experience descriptions / about /
    skill evidence / career summary must not crash the export."""
    ds_id = client.post("/datasets", files={"file": ("c.csv", FIXTURE_CSV.read_bytes(), "text/csv")}).json()["dataset"]["dataset_id"]
    client.post(f"/datasets/{ds_id}/enrich")

    from app import repositories as repo
    from app.models import ProfileSemantic, Skill

    people = repo.list_people(db, ds_id)
    p = people[0]
    p.about = "About \x00 text \x1f with \x0b control chars"
    for e in repo.get_experiences(db, p.id):
        e.description = (e.description or "") + " \x00\x01\x02 illegal \x0c tail"
    db.add(Skill(person_id=p.id, skill_name="X", skill_name_norm="x", source="llm_inference",
                 is_inferred=True, confidence=0.5, evidence="mentored \x0b juniors \x1f"))
    sem = ProfileSemantic(person_id=people[1].id, version=2, data={"career_summary": "Career \x00 summary \x1f"})
    db.add(sem)
    db.commit()

    r = client.get(f"/datasets/{ds_id}/export")
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert "Profiles" in wb.sheetnames and wb["Profiles"].max_row >= 2
